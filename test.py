# from openpyxl import Workbook
#
#
# wb = Workbook()
# ws = wb.active
#
# ws1 = wb.create_sheet("FirstSheet", 0)
# ws2 = wb.create_sheet("SecondSheet", 1)
# ws1.title = "First"
# ws.sheet_properties.tabColor = "AAAAAA"
# ws1.sheet_properties.tabColor = "FF0000"
#
# a = ws['B4']
# ws['B4'] = 4
# a.value = 210
# ws1['C8'] = 500
# colC = ws1['C']
# for cell in colC:
#     cell.value = cell.row
#
# wb.save('test.xlsx')


# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
#
# wb = Workbook()
dest_filename = 'empty_book.xlsx'
#
# ws1 = wb.active
# ws1.title = "range names"
#
# #wniosek: lista do row ładnie się wstawia po koljenych kolumnach
# for row in range(1, 4):
#     ws1.append(range(600))
#
# ws1['A6'] = "=SUM(B1:B3)"
#
# ws2 = wb.create_sheet(title="Pi")
# ws2['F5'] = 3.14
#
# ws3 = wb.create_sheet(title="Data")
#
# for row in range(10, 20):
#     for col in range(27, 54):
#         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
#
# ws1.merge_cells('A6:G6')
# ws1.column_dimensions.group('A', 'G', hidden=False) #taka zakładka
#
# wb.save(filename=dest_filename)

from openpyxl import load_workbook

wb = load_workbook(dest_filename)
print(wb.sheetnames)
ws = wb["Pi"]

from openpyxl.chart import BarChart, Reference, Series
# values = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=8)
# chart = BarChart()
# chart.type = "col"
# chart.style = 12
# # chart.x_axis.title = "Litery aaaaaaaaaaaa"
# # chart.y_axis.title = "Liczby"
# # chart.set_categories(Reference(ws, min_com=))
# chart.add_data(values, titles_from_data=True, from_rows=False)
# chart.shape = 4
# cats = Reference(ws, min_col=1, min_row=2, max_row=8)
# chart.set_categories(cats)
# ws.add_chart(chart, "E15")
# wb.save(dest_filename)

# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Bar Chart"
# chart1.y_axis.title = 'Test number'
# chart1.x_axis.title = 'Sample length (mm)'
# chart1.x_axis.tickLblPos = "low"
#
# data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
# cats = Reference(ws, min_col=1, min_row=2, max_row=7)
# chart1.add_data(data, titles_from_data=True)
# chart1.set_categories(cats)
# chart1.shape = 4
# ws.add_chart(chart1, "A10")
#
# from copy import deepcopy
#
# chart2 = deepcopy(chart1)
# chart2.style = 11
# chart2.type = "bar"
# chart2.title = "Horizontal Bar Chart"
#
# ws.add_chart(chart2, "G10")
#
#
# chart3 = deepcopy(chart1)
# chart3.type = "col"
# chart3.style = 12
# chart3.grouping = "stacked"
# chart3.overlap = 100
# chart3.title = 'Stacked Chart'
#
# ws.add_chart(chart3, "A27")
#
#
# chart4 = deepcopy(chart1)
# chart4.type = "bar"
# chart4.style = 13
# # chart4.grouping = "percentStacked"
# # chart4.overlap = 100
# chart4.title = 'Percent Stacked Chart'
#
# ws.add_chart(chart4, "G27")
# wb.save(dest_filename)


# import openpyxl module
import openpyxl

# import LineChart class from openpyxl.chart sub_module
from openpyxl.chart import LineChart, Reference

wb = openpyxl.Workbook()
sheet = wb.active

# write o to 9 in 1st column of the active sheet
for i in range(10):
    sheet.append([i])

values = Reference(sheet, min_col=1, min_row=1,
                   max_col=1, max_row=10)

# Create object of LineChart class
chart = LineChart()

chart.add_data(values)

# set the title of the chart
chart.title = " LINE-CHART "

# set the title of the x-axis
chart.x_axis.title = " X-AXIS "

# set the title of the y-axis
chart.y_axis.title = " Y-AXIS " ##X są dziwnie tytułowane, zobaczyć jak to działa na MS

# add chart to the sheet
# the top-left corner of a chart
# is anchored to cell E2 .
sheet.add_chart(chart, "E2")

# save the file
wb.save("LineChart.xlsx")

